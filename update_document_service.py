import os

with open('backend/app/services/document_service.py', 'w', encoding='utf-8') as f:
    f.write('''import os
import shutil
import uuid
import openpyxl
from openpyxl.utils import get_column_letter
import win32com.client
import pythoncom
from datetime import datetime

class DocumentService:
    @staticmethod
    def _convert_excel_to_pdf(excel_path: str, pdf_path: str):
        \"\"\"
        Convierte un archivo Excel a PDF utilizando win32com de forma invisible.
        \"\"\"
        pythoncom.CoInitialize()
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            wb = excel.Workbooks.Open(os.path.abspath(excel_path))
            # 0 es el formato xlTypePDF (en xlFixedFormatType)
            wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
            wb.Close(False)
        finally:
            excel.Quit()
            pythoncom.CoUninitialize()

    @staticmethod
    def generate_payslip(boleta_data: dict, output_format: str = "xlsx") -> str:
        \"\"\"
        Genera la Boleta de Pago en Excel (usando plantilla) y opcionalmente la convierte a PDF.
        \"\"\"
        template_path = os.path.join("backend", "app", "templates", "excel", "plantilla_boletas_de_pago.xlsx")
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Plantilla no encontrada en {template_path}")
            
        # Generar nombre temporal
        unique_id = uuid.uuid4().hex[:8]
        output_xlsx = f"boleta_{unique_id}.xlsx"
        output_pdf = f"boleta_{unique_id}.pdf"
        
        # Cargar excel
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Definir el mapeo exacto de celdas según la captura del usuario
        # B2: Código del Empleado (reemplazando "Codigo")
        ws['B2'] = boleta_data.get('internal_code', 'N/A')
        
        # C4 a D15
        ws['D4'] = boleta_data.get('empresa_nombre', '')
        ws['D5'] = boleta_data.get('empresa_nombre', '')
        ws['D6'] = boleta_data.get('nit', '')
        ws['D7'] = boleta_data.get('numero_patronal', '')
        ws['D8'] = f"{boleta_data.get('mes', '')}/{boleta_data.get('anio', '')}"
        ws['D9'] = boleta_data.get('ci', '')
        ws['D10'] = boleta_data.get('nombres', '')
        ws['D11'] = boleta_data.get('apellido_paterno', '')
        ws['D12'] = boleta_data.get('apellido_materno', '')
        ws['D13'] = boleta_data.get('fecha_ingreso', '')
        ws['D14'] = boleta_data.get('fecha_nacimiento', '')
        ws['D15'] = boleta_data.get('cargo', '')
        
        # Ingresos y Descuentos (D16 a D29)
        ws['D16'] = boleta_data.get('haber_basico', 0)
        ws['D17'] = boleta_data.get('bono_antiguedad', 0)
        ws['D18'] = 0  # Incremento
        ws['D19'] = boleta_data.get('haber_basico', 0) + boleta_data.get('bono_antiguedad', 0)  # Subtotal
        ws['D20'] = boleta_data.get('subsidio_natalidad', 0)
        ws['D21'] = boleta_data.get('aporte_gestora', 0)
        ws['D22'] = boleta_data.get('rc_iva', 0)
        ws['D24'] = boleta_data.get('otros_ingresos', 0)
        ws['D26'] = boleta_data.get('anticipos', 0)
        ws['D27'] = 0  # Desc Corporativas
        ws['D28'] = boleta_data.get('otros_descuentos', 0)
        ws['E28'] = boleta_data.get('total_ganado', 0)  # Mapeando TOTAL GANADO (en la captura parece usar E28 o D28)
        
        # Forzar un recálculo opcional o dejar que Excel lo haga
        
        wb.save(output_xlsx)
        wb.close()
        
        if output_format == "pdf":
            DocumentService._convert_excel_to_pdf(output_xlsx, output_pdf)
            os.remove(output_xlsx)
            return output_pdf
            
        return output_xlsx

    @staticmethod
    def generate_payroll_excel(payroll_data: list[dict], output_format: str = "xlsx") -> str:
        \"\"\"
        Genera un archivo Excel con la Planilla de Sueldos y opcionalmente a PDF.
        \"\"\"
        template_path = os.path.join("backend", "app", "templates", "excel", "plantilla_de_sueldos_y_salarios.xlsx")
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Plantilla no encontrada en {template_path}")
            
        unique_id = uuid.uuid4().hex[:8]
        output_xlsx = f"planilla_{unique_id}.xlsx"
        output_pdf = f"planilla_{unique_id}.pdf"
        
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Datos de cabecera generales (asumimos primera fila de datos)
        if payroll_data:
            empresa = payroll_data[0].get('empresa_nombre', '')
            nit = payroll_data[0].get('nit', '')
            nro_pat = payroll_data[0].get('numero_patronal', '')
            mes_anio = f"{payroll_data[0].get('mes', '')}/{payroll_data[0].get('anio', '')}"
            
            ws['C2'] = empresa
            ws['N2'] = nit
            ws['C3'] = nro_pat
            ws['V5'] = mes_anio  # Mes
        
        start_row = 9
        # Insertar filas necesarias manteniendo fórmulas (copiando la fila 9).
        # OpenPyxl no copia fórmulas automáticamente al insertar, así que las escribiremos 
        # o confiaremos en las fórmulas de Excel si las filas ya existen.
        # El usuario indicó que usemos las fórmulas del Excel, así que si faltan filas:
        for i, emp in enumerate(payroll_data):
            row = start_row + i
            # Si hay más de una fila en la plantilla, rellenamos. 
            # (Si openpyxl no evalúa fórmulas para el PDF, win32com sí lo hará al abrirlo!)
            
            ws.cell(row=row, column=1).value = i + 1
            ws.cell(row=row, column=2).value = emp.get('documento_identidad', '')
            ws.cell(row=row, column=3).value = f"{emp.get('apellido_paterno','')} {emp.get('apellido_materno', '')} {emp.get('nombres','')}".strip().replace("  ", " ").upper()
            ws.cell(row=row, column=4).value = emp.get('nacionalidad', 'Boliviana')
            ws.cell(row=row, column=5).value = emp.get('fecha_nacimiento', '')
            ws.cell(row=row, column=6).value = emp.get('sexo', '')
            ws.cell(row=row, column=7).value = emp.get('ocupacion', '')
            ws.cell(row=row, column=8).value = emp.get('fecha_ingreso', '')
            ws.cell(row=row, column=9).value = emp.get('horas_pagadas', 240)
            ws.cell(row=row, column=10).value = emp.get('dias_pagados', 30)
            ws.cell(row=row, column=11).value = float(emp.get('haber_basico', 0))
            ws.cell(row=row, column=12).value = float(emp.get('bono_antiguedad', 0))
            ws.cell(row=row, column=13).value = float(emp.get('bono_produccion', 0))
            ws.cell(row=row, column=14).value = float(emp.get('subsidio_frontera', 0))
            ws.cell(row=row, column=15).value = float(emp.get('trabajo_extraordinario', 0))
            ws.cell(row=row, column=16).value = float(emp.get('pago_dominical', 0))
            ws.cell(row=row, column=17).value = float(emp.get('otros_bonos', 0))
            ws.cell(row=row, column=18).value = float(emp.get('total_ganado', 0))
            ws.cell(row=row, column=19).value = float(emp.get('aporte_gestora', 0))
            ws.cell(row=row, column=20).value = float(emp.get('rc_iva', 0))
            ws.cell(row=row, column=21).value = float(emp.get('otros_descuentos', 0)) + float(emp.get('anticipos', 0))
            ws.cell(row=row, column=22).value = float(emp.get('total_descuentos', 0))
            ws.cell(row=row, column=23).value = float(emp.get('liquido_pagable', 0))
            
        wb.save(output_xlsx)
        wb.close()
        
        if output_format == "pdf":
            DocumentService._convert_excel_to_pdf(output_xlsx, output_pdf)
            os.remove(output_xlsx)
            return output_pdf
            
        return output_xlsx

    @staticmethod
    def generate_settlement_word(template_path: str, context: dict, output_path: str):
        from docxtpl import DocxTemplate
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"No se encontró la plantilla en {template_path}")
            
        doc = DocxTemplate(template_path)
        doc.render(context)
        doc.save(output_path)
        return output_path
''')