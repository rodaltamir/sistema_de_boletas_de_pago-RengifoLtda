import os
import shutil
import uuid
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

class DocumentService:

    @staticmethod
    def _set_cell_value(ws, coord, value):
        for merged_range in ws.merged_cells.ranges:
            if coord in merged_range:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                return
        ws[coord] = value

    @staticmethod
    def _numero_a_letras(numero: int) -> str:
        unidades = ["", "UN", "DOS", "TRES", "CUATRO", "CINCO", "SEIS", "SIETE", "OCHO", "NUEVE", "DIEZ", "ONCE", "DOCE", "TRECE", "CATORCE", "QUINCE", "DIECISEIS", "DIECISIETE", "DIECIOCHO", "DIECINUEVE", "VEINTE", "VEINTIUN", "VEINTIDOS", "VEINTITRES", "VEINTICUATRO", "VEINTICINCO", "VEINTISEIS", "VEINTISIETE", "VEINTIOCHO", "VEINTINUEVE"]
        decenas = ["", "DIEZ", "VEINTE", "TREINTA", "CUARENTA", "CINCUENTA", "SESENTA", "SETENTA", "OCHENTA", "NOVENTA"]
        centenas = ["", "CIENTO", "DOSCIENTOS", "TRESCIENTOS", "CUATROCIENTOS", "QUINIENTOS", "SEISCIENTOS", "SETECIENTOS", "OCHOCIENTOS", "NOVECIENTOS"]
        
        if numero == 0: return "CERO"
        if numero <= 29: return unidades[numero]
        if numero <= 99:
            return decenas[numero // 10] if numero % 10 == 0 else decenas[numero // 10] + " Y " + unidades[numero % 10]
        if numero == 100: return "CIEN"
        if numero <= 999:
            return centenas[numero // 100] if numero % 100 == 0 else centenas[numero // 100] + " " + DocumentService._numero_a_letras(numero % 100)
        if numero == 1000: return "MIL"
        if numero <= 999999:
            miles = numero // 1000
            resto = numero % 1000
            str_miles = "MIL" if miles == 1 else DocumentService._numero_a_letras(miles) + " MIL"
            return str_miles if resto == 0 else str_miles + " " + DocumentService._numero_a_letras(resto)
        if numero == 1000000: return "UN MILLON"
        if numero <= 999999999:
            millones = numero // 1000000
            resto = numero % 1000000
            str_millones = "UN MILLON" if millones == 1 else DocumentService._numero_a_letras(millones) + " MILLONES"
            return str_millones if resto == 0 else str_millones + " " + DocumentService._numero_a_letras(resto)
        
        return str(numero)

    @staticmethod
    def _convert_excel_to_pdf(excel_path: str, pdf_path: str):
        """
        Convierte un archivo Excel a PDF. En Windows usa win32com, en Linux usa LibreOffice.
        """
        if os.name == 'nt':
            import win32com.client
            import pythoncom
            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            try:
                wb = excel.Workbooks.Open(os.path.abspath(excel_path))
                wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
                wb.Close(False)
            finally:
                excel.Quit()
                pythoncom.CoUninitialize()
        else:
            import subprocess
            # LibreOffice headless mode
            outdir = os.path.dirname(os.path.abspath(pdf_path))
            subprocess.run([
                'libreoffice', '--headless', '--convert-to', 'pdf',
                '--outdir', outdir, os.path.abspath(excel_path)
            ], check=True)
            
            # LibreOffice generates the file with the same base name but .pdf extension
            base_name = os.path.splitext(os.path.basename(excel_path))[0]
            generated_pdf = os.path.join(outdir, f"{base_name}.pdf")
            
            # Rename it to the target pdf_path if it differs
            if generated_pdf != os.path.abspath(pdf_path):
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
                os.rename(generated_pdf, pdf_path)

    @staticmethod
    def generate_payslip(boleta_data: dict, output_format: str = "xlsx") -> str:
        """
        Genera la Boleta de Pago en Excel (usando plantilla) y opcionalmente la convierte a PDF.
        """
        template_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "templates", "excel", "plantilla_boletas_de_pago.xlsx"))
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Plantilla no encontrada en {template_path}")
            
        # Generar nombre temporal
        unique_id = uuid.uuid4().hex[:8]
        output_xlsx = f"boleta_{unique_id}.xlsx"
        output_pdf = f"boleta_{unique_id}.pdf"
        
        # Cargar excel
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        internal_code = boleta_data.get('internal_code', '')
        if isinstance(internal_code, str):
            internal_code = internal_code.replace('"', '').replace("'", "")

        # B2: Empresa Name
        DocumentService._set_cell_value(ws, 'B2', boleta_data.get('empresa_nombre', '').upper())
        DocumentService._set_cell_value(ws, 'D2', "Nº:")
        DocumentService._set_cell_value(ws, 'E2', internal_code)
        
        DocumentService._set_cell_value(ws, 'B3', f"Nro. Patronal: {boleta_data.get('numero_patronal', '')}")
        
        mes_int = int(boleta_data.get('mes', 0))
        MESES = {1:"Enero", 2:"Febrero", 3:"Marzo", 4:"Abril", 5:"Mayo", 6:"Junio", 7:"Julio", 8:"Agosto", 9:"Septiembre", 10:"Octubre", 11:"Noviembre", 12:"Diciembre"}
        mes_nombre = MESES.get(mes_int, str(mes_int))
        anio = boleta_data.get('anio', '')
        DocumentService._set_cell_value(ws, 'B7', f"MES {mes_nombre}")
        DocumentService._set_cell_value(ws, 'C7', f"AÑO {anio}")
        
        # Fecha fin de mes (asumiendo 30 o 31)
        fecha_fin = f"30/{mes_int:02d}/{anio}" if mes_int in [4,6,9,11] else f"31/{mes_int:02d}/{anio}"
        if mes_int == 2: fecha_fin = f"28/02/{anio}"
        DocumentService._set_cell_value(ws, 'D7', f"FECHA {fecha_fin}")
        
        DocumentService._set_cell_value(ws, 'B9', f"CODIGO :      {internal_code}")
        
        emp_nombres = boleta_data.get('nombres') or ''
        emp_pat = boleta_data.get('apellido_paterno') or ''
        emp_mat = boleta_data.get('apellido_materno') or ''
        full_emp_name = f"{emp_pat} {emp_mat} {emp_nombres}".strip().replace("  ", " ").upper()
        DocumentService._set_cell_value(ws, 'C9', f"NOMBRE : {full_emp_name}")
        
        DocumentService._set_cell_value(ws, 'B10', f"CARGO : {boleta_data.get('cargo', '').upper()}")
        DocumentService._set_cell_value(ws, 'B11', f"FECHA INGRESO : {boleta_data.get('fecha_ingreso', '')}")
        DocumentService._set_cell_value(ws, 'D11', "SALDO I.V.A. :     0.00")
        
        # Ingresos y Descuentos
        hb = float(boleta_data.get('haber_basico', 0) or 0)
        ba = float(boleta_data.get('bono_antiguedad', 0) or 0)
        sub_nat = float(boleta_data.get('subsidio_natalidad', 0) or 0)
        otros_ing = float(boleta_data.get('otros_ingresos', 0) or 0)
        total_ganado = float(boleta_data.get('total_ganado', 0) or 0)
        
        gestora = float(boleta_data.get('aporte_gestora', 0) or 0)
        rc_iva = float(boleta_data.get('rc_iva', 0) or 0)
        anticipos = float(boleta_data.get('anticipos', 0) or 0)
        otros_des = float(boleta_data.get('otros_descuentos', 0) or 0)
        total_descuentos = float(boleta_data.get('total_descuentos', gestora + rc_iva + anticipos + otros_des) or 0)
        liquido = float(boleta_data.get('liquido_pagable', total_ganado - total_descuentos) or 0)

        # Ingresos left side
        DocumentService._set_cell_value(ws, 'B13', "Sueldo Básico")
        DocumentService._set_cell_value(ws, 'C13', hb)
        DocumentService._set_cell_value(ws, 'B14', "Bono de Antigüedad")
        DocumentService._set_cell_value(ws, 'C14', ba)
        DocumentService._set_cell_value(ws, 'B15', "Subsidio de Natalidad")
        DocumentService._set_cell_value(ws, 'C15', sub_nat)
        DocumentService._set_cell_value(ws, 'B16', "Otros Ingresos")
        DocumentService._set_cell_value(ws, 'C16', otros_ing)
        
        # Descuentos right side
        DocumentService._set_cell_value(ws, 'D13', "Aporte Gestora")
        DocumentService._set_cell_value(ws, 'E13', gestora)
        DocumentService._set_cell_value(ws, 'D14', "R.C. - I.V.A.")
        DocumentService._set_cell_value(ws, 'E14', rc_iva)
        DocumentService._set_cell_value(ws, 'D15', "Anticipos")
        DocumentService._set_cell_value(ws, 'E15', anticipos)
        DocumentService._set_cell_value(ws, 'D16', "Otros Descuentos")
        DocumentService._set_cell_value(ws, 'E16', otros_des)

        # Totals
        DocumentService._set_cell_value(ws, 'C18', total_ganado)
        DocumentService._set_cell_value(ws, 'E18', total_descuentos)
        
        # Liquido Pagable
        # Descombinamos B21:C21 si están combinados para evitar sobreescribir el texto
        try:
            ws.unmerge_cells('B21:C21')
        except:
            pass
            
        ws['B21'] = "LIQUIDO PAGABLE:"
        ws['B21'].font = openpyxl.styles.Font(name="Times New Roman", size=11, bold=True)
        ws['C21'] = liquido
        ws['C21'].font = openpyxl.styles.Font(name="Times New Roman", size=13, bold=True)
        ws['C21'].number_format = '#,##0.00'
        ws['C21'].alignment = openpyxl.styles.Alignment(horizontal='right')
        
        entero = int(liquido)
        decimal = int(round((liquido - entero) * 100))
        literal_entero = DocumentService._numero_a_letras(entero)
        
        try:
            ws.merge_cells('D21:E21')
        except:
            pass
        ws['D21'] = f"(Son: {literal_entero} con {decimal}/100 Bolivianos)"
        ws['D21'].font = openpyxl.styles.Font(name="Times New Roman", size=9, italic=True)
        ws['D21'].alignment = openpyxl.styles.Alignment(shrink_to_fit=True, vertical='center', horizontal='left')
        
        # Signatures
        DocumentService._set_cell_value(ws, 'B25', "...........................................................")
        DocumentService._set_cell_value(ws, 'B26', "Verificado Contabilidad/Gerencia")
        
        DocumentService._set_cell_value(ws, 'D25', "...........................................................")
        DocumentService._set_cell_value(ws, 'D26', full_emp_name)
        
        wb.save(output_xlsx)
        wb.close()
        
        if output_format == "pdf":
            DocumentService._convert_excel_to_pdf(output_xlsx, output_pdf)
            os.remove(output_xlsx)
            return output_pdf
            
        return output_xlsx

    @staticmethod
    def generate_payroll_excel(payroll_data: list[dict], output_format: str = "xlsx") -> str:
        """
        Genera un archivo Excel con la Planilla de Sueldos y opcionalmente a PDF.
        """
        template_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "templates", "excel", "plantilla_de_sueldos_y_salarios.xlsx"))
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Plantilla no encontrada en {template_path}")
            
        unique_id = uuid.uuid4().hex[:8]
        output_xlsx = f"planilla_{unique_id}.xlsx"
        output_pdf = f"planilla_{unique_id}.pdf"
        
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        if payroll_data:
            empresa = payroll_data[0].get('empresa_nombre', '')
            nit = payroll_data[0].get('nit', '')
            patronal = payroll_data[0].get('numero_patronal', '')
            mes_int = int(payroll_data[0].get('mes', 0))
            MESES = {1:"ENERO", 2:"FEBRERO", 3:"MARZO", 4:"ABRIL", 5:"MAYO", 6:"JUNIO", 7:"JULIO", 8:"AGOSTO", 9:"SEPTIEMBRE", 10:"OCTUBRE", 11:"NOVIEMBRE", 12:"DICIEMBRE"}
            anio = payroll_data[0].get('anio', '')
            
            DocumentService._set_cell_value(ws, 'C2', empresa.upper())
            try:
                ws['C2'].font = openpyxl.styles.Font(name="Arial", size=10, bold=True)
                ws['C2'].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
            except: pass

            DocumentService._set_cell_value(ws, 'F3', nit)
            try:
                ws['F3'].font = openpyxl.styles.Font(name="Arial", size=10, bold=True)
                ws['F3'].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
            except: pass

            DocumentService._set_cell_value(ws, 'O2', nit)
            try:
                ws['O2'].font = openpyxl.styles.Font(name="Arial", size=10, bold=True)
                ws['O2'].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
            except: pass

            DocumentService._set_cell_value(ws, 'O3', patronal)
            try:
                ws['O3'].font = openpyxl.styles.Font(name="Arial", size=10, bold=True)
                ws['O3'].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
            except: pass
            
            # Recuperar texto original si existe o poner default
            base_text = "CORRESPONDIENTE AL MES DE"
            DocumentService._set_cell_value(ws, 'U6', f"{base_text} {MESES.get(mes_int, str(mes_int))} DE {anio}")
            try:
                ws['U6'].font = openpyxl.styles.Font(name="Arial", size=11, bold=True)
                ws['U6'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            except: pass
            
            # Paginacion
            DocumentService._set_cell_value(ws, 'V2', 1)
            DocumentService._set_cell_value(ws, 'X2', 1)
        
        start_row = 11
        num_employees = len(payroll_data)
        
        if num_employees > 2:
            ws.insert_rows(13, amount=num_employees - 2)
            
        for i, emp in enumerate(payroll_data):
            row = start_row + i
            # Para las filas normales de empleados, evitamos el helper custom si da problemas
            # pero sabemos que A-W estan descombinadas o sus celdas principales coinciden
            try: ws.cell(row=row, column=1).value = i + 1
            except: pass
            try: ws.cell(row=row, column=2).value = emp.get('documento_identidad', '')
            except: pass
            try: ws.cell(row=row, column=3).value = f"{emp.get('apellido_paterno','')} {emp.get('apellido_materno', '')} {emp.get('nombres','')}".strip().replace("  ", " ").upper()
            except: pass
            try: ws.cell(row=row, column=4).value = emp.get('nacionalidad', 'Boliviana')
            except: pass
            try: ws.cell(row=row, column=5).value = emp.get('fecha_nacimiento', '')
            except: pass
            try: ws.cell(row=row, column=6).value = emp.get('sexo', '')
            except: pass
            try: ws.cell(row=row, column=7).value = emp.get('ocupacion', '')
            except: pass
            try: ws.cell(row=row, column=8).value = emp.get('fecha_ingreso', '')
            except: pass
            try: ws.cell(row=row, column=9).value = emp.get('horas_pagadas', 240)
            except: pass
            try: ws.cell(row=row, column=10).value = emp.get('dias_pagados', 30)
            except: pass
            try: ws.cell(row=row, column=11).value = float(emp.get('haber_basico', 0))
            except: pass
            try: ws.cell(row=row, column=12).value = float(emp.get('bono_antiguedad', 0))
            except: pass
            try: ws.cell(row=row, column=13).value = float(emp.get('bono_produccion', 0))
            except: pass
            try: ws.cell(row=row, column=14).value = float(emp.get('subsidio_frontera', 0))
            except: pass
            try: ws.cell(row=row, column=15).value = float(emp.get('trabajo_extraordinario', 0))
            except: pass
            try: ws.cell(row=row, column=16).value = float(emp.get('pago_dominical', 0))
            except: pass
            try: ws.cell(row=row, column=17).value = float(emp.get('otros_bonos', 0))
            except: pass
            try: ws.cell(row=row, column=18).value = float(emp.get('total_ganado', 0))
            except: pass
            try: ws.cell(row=row, column=19).value = float(emp.get('aporte_gestora', 0))
            except: pass
            try: ws.cell(row=row, column=20).value = float(emp.get('rc_iva', 0))
            except: pass
            otros = float(emp.get('otros_descuentos', 0)) + float(emp.get('anticipos', 0))
            try: ws.cell(row=row, column=21).value = otros
            except: pass
            try: ws.cell(row=row, column=22).value = float(emp.get('total_descuentos', 0))
            except: pass
            try: ws.cell(row=row, column=23).value = float(emp.get('liquido_pagable', 0))
            except: pass
            
            # Formatting para los números en la tabla
            for col in range(11, 24):
                try: ws.cell(row=row, column=col).number_format = '#,##0.00'
                except: pass

        # Fila TOTALES
        totales_row = start_row + max(num_employees, 2)
        try:
            ws.cell(row=totales_row, column=1).value = 'TOTALES'
            ws.cell(row=totales_row, column=1).font = openpyxl.styles.Font(name="Arial", size=10, bold=True)
        except: pass
        
        for col in range(11, 24):
            col_letter = get_column_letter(col)
            try:
                c = ws.cell(row=totales_row, column=col)
                c.value = f"=SUM({col_letter}{start_row}:{col_letter}{totales_row-1})"
                c.font = openpyxl.styles.Font(name="Arial", size=10, bold=True)
                c.number_format = '#,##0.00'
            except: pass
            
        # Fila EMPLEADOR
        empleador_row = totales_row + 4
        if payroll_data:
            emp_nombres = payroll_data[0].get('empleador_nombres') or ''
            emp_pat = payroll_data[0].get('empleador_apellido_paterno') or ''
            emp_mat = payroll_data[0].get('empleador_apellido_materno') or ''
            emp_ci = payroll_data[0].get('empleador_ci') or ''
            
            full_emp_name = f"{emp_pat} {emp_mat} {emp_nombres}".strip().replace("  ", " ").upper()
            
            # Nombre del Empleador (Centrado sin lineas de puntos)
            try:
                c1 = ws.cell(row=empleador_row, column=4)
                c1.value = full_emp_name
                c1.font = openpyxl.styles.Font(name="Arial", size=10, bold=True)
                c1.alignment = openpyxl.styles.Alignment(horizontal='center')
            except: pass
            
            # CI del Empleador
            try:
                c2 = ws.cell(row=empleador_row, column=12)
                c2.value = f"CI: {emp_ci}"
                c2.font = openpyxl.styles.Font(name="Arial", size=10, bold=True)
                c2.alignment = openpyxl.styles.Alignment(horizontal='center')
            except: pass

        wb.save(output_xlsx)
        wb.close()
        
        if output_format == "pdf":
            DocumentService._convert_excel_to_pdf(output_xlsx, output_pdf)
            os.remove(output_xlsx)
            return output_pdf
            
        return output_xlsx

    @staticmethod
    def generate_prefiniquito_excel(data: dict, output_format: str = "xlsx") -> str:
        """
        Genera un archivo Excel con la preliquidación basado en la plantilla y opcionalmente lo exporta a PDF.
        """
        template_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "templates", "excel", "plantilla_prefiniquitos.xlsx"))
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Plantilla no encontrada en {template_path}")
            
        unique_id = uuid.uuid4().hex[:8]
        output_xlsx = f"prefiniquito_{unique_id}.xlsx"
        output_pdf = f"prefiniquito_{unique_id}.pdf"
        
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        format_bs = lambda x: f"{float(x):,.2f} Bs".replace(",", "X").replace(".", ",").replace("X", ".") if x is not None else "0,00 Bs"
        
        # Helper para rellenar
        def set_val(cell, val):
            if val is not None:
                ws[cell] = val
                try: ws[cell].alignment = openpyxl.styles.Alignment(horizontal='right')
                except: pass
                
        # Llenar datos principales
        ws['B7'] = f"NOMBRE DEL TRABAJADOR:     {data.get('nombre_trabajador', '').upper()}"
        ws['B8'] = f"RAZÓN SOCIAL DEL EMPLEADOR:  {data.get('razon_social', '').upper()}"
        ws['B9'] = f"FECHA DE INGRESO:          {data.get('fecha_ingreso', '')}"
        ws['B10'] = f"FECHA DE RETIRO:           {data.get('fecha_retiro', '')}"
        
        anios = data.get('anios_trabajados', 0)
        meses = data.get('meses_trabajados', 0)
        dias = data.get('dias_trabajados', 0)
        ws['B11'] = f"TIEMPO DE TRABAJO:         Años: {anios}       Meses: {meses}      Días: {dias}"
        
        ws['B12'] = f"BONO DE ANTIGUEDAD:        0,00 Bs"
        
        # Sueldo promedio formateado
        sp = data.get('sueldo_promedio', 0)
        sp_str = f"{float(sp):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        ws['B13'] = f"SUELDO PROMEDIO (Bs.):     {sp_str}"
        
        # Desahucio
        set_val('F16', format_bs(data.get('desahucio', 0)))
        
        # Indemnizacion
        if anios > 0:
            ws['D18'] = f"{anios} Años"
            set_val('F18', format_bs(data.get('indemnizacion_anios', 0)))
            ws.row_dimensions[18].hidden = False
        else:
            ws['D18'] = ""
            set_val('F18', "")
            ws.row_dimensions[18].hidden = True
            
        if meses > 0:
            ws['D19'] = f"{meses} Meses"
            set_val('F19', format_bs(data.get('indemnizacion_meses', 0)))
            ws.row_dimensions[19].hidden = False
        else:
            ws['D19'] = ""
            set_val('F19', "")
            ws.row_dimensions[19].hidden = True
            
        if dias > 0:
            ws['D20'] = f"{dias} Días"
            set_val('F20', format_bs(data.get('indemnizacion_dias', 0)))
            ws.row_dimensions[20].hidden = False
        else:
            ws['D20'] = ""
            set_val('F20', "")
            ws.row_dimensions[20].hidden = True
        
        # Aguinaldo (Calcular tiempo en meses y días si no vienen dados explícitamente, o simplemente usar los counts)
        ag_meses_count = data.get('aguinaldo_meses_count', 0)
        ag_dias_count = data.get('aguinaldo_dias_count', 0)
        
        if ag_meses_count > 0:
            ws['D22'] = f"{ag_meses_count} Meses"
            set_val('F22', format_bs(data.get('aguinaldo_meses', 0)))
            ws.row_dimensions[22].hidden = False
        else:
            ws['D22'] = ""
            set_val('F22', "")
            ws.row_dimensions[22].hidden = True
            
        if ag_dias_count > 0:
            ws['D23'] = f"{ag_dias_count} Días"
            set_val('F23', format_bs(data.get('aguinaldo_dias', 0)))
            ws.row_dimensions[23].hidden = False
        else:
            ws['D23'] = ""
            set_val('F23', "")
            ws.row_dimensions[23].hidden = True
        
        # Vacaciones
        vac_dias = data.get('dias_vacacion_pendientes', 0)
        ws['D25'] = f"{vac_dias} Días"
        set_val('F25', format_bs(data.get('vacaciones', 0)))
        
        # Otros
        ws['D27'] = "OTROS PAGOS"
        set_val('F27', format_bs(data.get('otros_pagos', 0)))
        set_val('F28', "0,00 Bs") # Línea extra en otros
        
        # Descuentos
        set_val('F30', format_bs(data.get('descuentos', 0)))
        
        # Totales
        set_val('F31', format_bs(data.get('total_calculo', 0)))
        set_val('F32', format_bs(data.get('multa_30', 0)))
        set_val('F33', format_bs(data.get('total_final', 0)))

        wb.save(output_xlsx)
        wb.close()
        
        if output_format == "pdf":
            DocumentService._convert_excel_to_pdf(output_xlsx, output_pdf)
            os.remove(output_xlsx)
            return output_pdf
            
        return output_xlsx

    @staticmethod
    def generate_settlement_word(template_path: str, context: dict, output_path: str):
        from docxtpl import DocxTemplate
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"No se encontró la plantilla en {template_path}")
            
        doc = DocxTemplate(template_path)
        doc.render(context)
        doc.save(output_path)
        return output_path
